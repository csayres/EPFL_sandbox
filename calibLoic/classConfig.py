#cython: language_level=3
import os
import glob
from scipy import io
import shutil
import time
import csv
from scipy import io
import json
import logger as log
import openpyxl
import miscmath as mm
import numpy as np
import DEFINES
import errors

class Config:
	__slots__ = (	'currentProjectTime',\
					'generalProjectFolder',\
					'currentProjectFolder',\
					'resultFolder',\
					'positionerFolderPrefix',\
					'positionerFolderSuffix',\
					'lifetimeSuffix',\
					'figureFolder',\
					'overviewsFolder',\
					'figureExtension',\
					'overviewExtension',\
					'resultsOverviewFile',\
					'resultsOverviewAutosave',\
					'positionerModelFile',\
					'positionerModelExtension',\
					'generalConfigFolder',\
					'testBenchFolder',\
					'cameraFolder',\
					'positionersFolder',\
					'requirementsFolder',\
					'calibrationsFolder',\
					'testsFolder',\
					'configFolder',\
					'resultsOverviewTemplateFile',\
					'testBenchFileExtension',\
					'cameraFileExtension',\
					'positionersFileExtension',\
					'requirementsFileExtension',\
					'calibrationsFileExtension',\
					'testsFileExtension',\
					'currentTestBenchFile',\
					'currentPositionerFile',\
					'currentRequirementsFile',\
					'currentFastCalibrationFile',\
					'currentCalibrationFile',\
					'currentTestFile',\
					'configFile',\
					'configFileExtension',\
					'calibrationResultsFile',\
					'testResultsFile',\
					'calibrationResultsFileExt',\
					'testResultsFileExt',\
					'lifetimeIterationFolderName',\
					'resultsLoadingFolder',\
					'preloadPositionerModel',\
					'calibrateDatum',\
					'calibrateMotor',\
					'calibrateCogging',\
					'IDsToLoad',\
					'preheatBenchTime',\
					'preheatBench',\
					'doFastCalibRun',\
					'doCalibRun',\
					'overwritePositionerModel',\
					'loadCalibRun',\
					'doTestRun',\
					'loadTestRun',\
					'nbTestingLoops',\
					'currentLifetimeIteration',\
					'reloadCalibParEachIter',\
					'reloadTestParEachIter',\
					'doLivePlot',\
					'plotResults',\
					'saveInQc',\
					'sendMail',\
					'mailReceivers')

	def __init__(self):
		self.currentProjectTime					= time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime(time.time()))

		#project parameters
		self.generalProjectFolder 				= 'Projects'			#generalProjectFolder
		self.currentProjectFolder				= 'Blackbird'			#generalProjectFolder\currentProjectFolder
		self.resultFolder						= 'All_calibrations'	#generalProjectFolder\currentProjectFolder\resultFolder
		self.positionerFolderPrefix				= 'Positioner'			#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\
		self.positionerFolderSuffix				= ''					#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix
		self.lifetimeSuffix 					= 'lifetime'
		self.figureFolder						= 'Figures'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\figureFolder
		self.overviewsFolder					= 'Overview'			#generalProjectFolder\currentProjectFolder\overviewsFolder
		self.figureExtension					= '.png'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\figureFolder\*figureExtension
		self.overviewExtension					= '.png'				#generalProjectFolder\currentProjectFolder\overviewsFolder\*overviewExtension
		self.resultsOverviewFile				= 'Results.xlsx'		#generalProjectFolder\currentProjectFolder\resultsOverviewFile
		self.resultsOverviewAutosave 			= 'Results_autosave.xlsx'
		self.positionerModelFile				= 'Model'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\positionerModelFile+positionerID
		self.positionerModelExtension			= '.json'				#generalProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\positionerModelFile+positionerID+positionerModelExtension


		#configuration files
		self.generalConfigFolder				= 'Config'				#generalConfigFolder
		self.testBenchFolder					= 'TestBenches'			#generalConfigFolder\testBenchFolder
		self.cameraFolder						= 'Cameras'				#generalConfigFolder\cameraFolder
		self.positionersFolder					= 'Positioners'			#generalConfigFolder\positionersFolder
		self.requirementsFolder					= 'Requirements'		#generalConfigFolder\requirementsFolder
		self.calibrationsFolder					= 'Calibrations'		#generalConfigFolder\calibrationsFolder
		self.testsFolder						= 'Tests'				#generalConfigFolder\testsFolder
		self.configFolder 						= 'General'				#generalConfigFolder\configFolder
		self.resultsOverviewTemplateFile 		= 'Results Template.xlsx'#generalConfigFolder\configFolder\resultsOverviewTemplateFile
		self.testBenchFileExtension 			= '.json'				#generalConfigFolder\testBenchFolder\*testbenchFileExtension
		self.cameraFileExtension 				= '.mat'				#generalConfigFolder\cameraFolder\*cameraFileExtension
		self.positionersFileExtension			= '.json'				#generalConfigFolder\*positionersFileExtension
		self.requirementsFileExtension			= '.json'				#generalConfigFolder\*requirementsFileExtension
		self.calibrationsFileExtension			= '.json'				#generalConfigFolder\calibrationsFolder\*calibrationsFileExtension
		self.testsFileExtension 				= '.json'				#generalConfigFolder\testsFolder\*testsFileExtension
		self.currentTestBenchFile				= ''					#generalConfigFolder\testBenchFolder\currentTestbenchFile
		self.currentPositionerFile				= ''					#generalConfigFolder\testBenchFolder\currentPositionerFile
		self.currentRequirementsFile			= ''					#generalConfigFolder\testBenchFolder\currentRequirementsFile
		self.currentFastCalibrationFile			= ''					#generalConfigFolder\calibrationsFolder\currentFastCalibrationFile
		self.currentCalibrationFile				= ''					#generalConfigFolder\calibrationsFolder\currentCalibrationFile
		self.currentTestFile 					= ''					#generalConfigFolder\testsFolder\currentTestFile
		self.configFile 						= 'config'				#generalConfigFolder\configFolder\configFile
		self.configFileExtension				= '.json'				#generalConfigFolder\configFolder\configFile+configFileExtension

		#testing files output
		self.calibrationResultsFile				= 'calibResults'		#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\calibrationResultsFile
		self.testResultsFile					= 'testResults'			#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\testResultsFile
		self.calibrationResultsFileExt 			= '.json'				#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\calibrationResultsFile+fileID+calibrationResultsFileExt
		self.testResultsFileExt 				= '.json'				#GeneralProjectFolder\currentProjectFolder\resultFolder\positionerFolderPrefix+positionerID\currentProjectTime+positionerFolderSuffix\testResultsFile+fileID+testResultsFileExt
		self.lifetimeIterationFolderName 		= 'Iteration' 		
		self.resultsLoadingFolder 				= DEFINES.CONFIG_LOAD_LATEST_RESULT	 				#projectTime+folderSuffix DEFINES.CONFIG_LOAD_LATEST_RESULT
		self.IDsToLoad 							= []
		
		#program parameters
		self.preloadPositionerModel				= False

		self.calibrateMotor 					= False
		self.calibrateDatum 					= False
		self.calibrateCogging 					= False

		self.preheatBenchTime 					= DEFINES.CONFIG_PREHEAT_BENCH_DEFAULT_TIME
		self.preheatBench 						= True

		self.doFastCalibRun						= True
		self.doCalibRun							= True
		self.overwritePositionerModel			= True
		self.loadCalibRun 						= False
		self.doTestRun							= True
		self.loadTestRun 						= False
		self.nbTestingLoops						= 1
		self.currentLifetimeIteration 			= 0

		self.reloadCalibParEachIter 			= False
		self.reloadTestParEachIter 				= False

		self.doLivePlot							= False
		self.sendMail 							= True
		self.plotResults 						= True
		self.saveInQc 							= True
		self.mailReceivers 						= ['loic.grossen@epfl.ch']#,'luzius.kronig@epfl.ch','ricardo.araujo@epfl.ch']

	def load(self,fileName):
		#Load all the data in the file, exculding the fileInfos
		try:
			with open(os.path.join(fileName),'r') as inFile:
				variablesToLoad=json.load(inFile)
				for key in variablesToLoad.keys():
					if key in type(self).__slots__:
						setattr(self, key, variablesToLoad[key])
					else:
						log.message(DEFINES.LOG_MESSAGE_PRIORITY_DEBUG_WARNING,1,f'Unexpected data was encountered during the loading of the general parameters. Faulty key: {key}')
						if DEFINES.RAISE_ERROR_ON_UNEXPECTED_KEY:
							raise errors.IOError('Unexpected data was encountered during the loading of the general parameters') from None
						
		except OSError:
			raise errors.IOError('The general parameters file could not be found') from None

	def save(self):
		filePath = os.path.join(self.generalConfigFolder,self.configFolder)
		fileName = self.configFile+self.configFileExtension
		variablesToSave = {}

		variablesToSave['currentProjectFolder']			= self.currentProjectFolder
		variablesToSave['positionerFolderSuffix']		= self.positionerFolderSuffix
		variablesToSave['currentTestBenchFile']			= self.currentTestBenchFile
		variablesToSave['currentPositionerFile']		= self.currentPositionerFile
		variablesToSave['currentRequirementsFile']		= self.currentRequirementsFile
		variablesToSave['currentFastCalibrationFile']	= self.currentFastCalibrationFile
		variablesToSave['currentCalibrationFile']		= self.currentCalibrationFile
		variablesToSave['currentTestFile']				= self.currentTestFile
		variablesToSave['configFile']					= self.configFile
		variablesToSave['resultsLoadingFolder'] 		= self.resultsLoadingFolder
		variablesToSave['IDsToLoad'] 					= self.IDsToLoad
		variablesToSave['preloadPositionerModel']		= self.preloadPositionerModel
		variablesToSave['calibrateMotor'] 				= self.calibrateMotor
		variablesToSave['calibrateDatum'] 				= self.calibrateDatum
		variablesToSave['calibrateCogging']				= self.calibrateCogging
		variablesToSave['preheatBenchTime'] 			= self.preheatBenchTime
		variablesToSave['preheatBench'] 				= self.preheatBench
		variablesToSave['doFastCalibRun']				= self.doFastCalibRun
		variablesToSave['doCalibRun']					= self.doCalibRun
		variablesToSave['overwritePositionerModel']		= self.overwritePositionerModel
		variablesToSave['loadCalibRun']					= self.loadCalibRun
		variablesToSave['doTestRun']					= self.doTestRun
		variablesToSave['loadTestRun']					= self.loadTestRun
		variablesToSave['nbTestingLoops']				= self.nbTestingLoops
		variablesToSave['reloadCalibParEachIter']		= self.reloadCalibParEachIter
		variablesToSave['reloadTestParEachIter']		= self.reloadTestParEachIter
		variablesToSave['doLivePlot']					= self.doLivePlot
		variablesToSave['plotResults'] 					= self.plotResults
		variablesToSave['saveInQc'] 					= self.saveInQc
		variablesToSave['sendMail']						= self.sendMail
		variablesToSave['mailReceivers']				= self.mailReceivers

		os.makedirs(filePath, exist_ok=True)
		with open(os.path.join(filePath, fileName),'w+') as outFile:
			json.dump(variablesToSave, outFile, separators = (',\n',': '))

	def reset_project_time(self):
		self.currentProjectTime	= time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime(time.time()))

	def get_camera_path(self):
		return os.path.join(self.generalConfigFolder,self.cameraFolder)

	def get_testbench_path(self):
		return os.path.join(self.generalConfigFolder,self.testBenchFolder)

	def get_fast_calib_param_path(self):
		return os.path.join(self.generalConfigFolder, self.calibrationsFolder)

	def get_calib_param_path(self):
		return os.path.join(self.generalConfigFolder, self.calibrationsFolder)
	
	def get_positioner_physics_path(self):
		return os.path.join(self.generalConfigFolder, self.positionersFolder)

	def get_positioner_requirements_path(self):
		return os.path.join(self.generalConfigFolder, self.requirementsFolder)

	def get_test_param_path(self):
		return os.path.join(self.generalConfigFolder, self.testsFolder)

	def get_config_fileName(self):
		return os.path.join(self.generalConfigFolder, self.configFolder, self.configFile + self.configFileExtension)

	def get_current_testBench_fileName(self):
		return os.path.join(self.generalConfigFolder, self.testBenchFolder, self.currentTestBenchFile + self.testBenchFileExtension)

	def get_current_fast_calib_param_fileName(self):
		return os.path.join(self.generalConfigFolder, self.calibrationsFolder, self.currentFastCalibrationFile+self.calibrationsFileExtension)

	def get_current_calib_param_fileName(self):
		return os.path.join(self.generalConfigFolder, self.calibrationsFolder, self.currentCalibrationFile+self.calibrationsFileExtension)
	
	def get_current_positioner_physics_fileName(self):
		return os.path.join(self.generalConfigFolder, self.positionersFolder, self.currentPositionerFile+self.positionersFileExtension)

	def get_current_positioner_requirements_fileName(self):
		return os.path.join(self.generalConfigFolder, self.requirementsFolder, self.currentRequirementsFile+self.requirementsFileExtension)

	def get_current_test_param_fileName(self):
		return os.path.join(self.generalConfigFolder, self.testsFolder, self.currentTestFile+self.testsFileExtension)

	def get_all_testbench_filenames(self):
		filenames = []

		for file in os.listdir(self.get_testbench_path()):
			if file.endswith(".txt"):
				filenames.append(file)

		return filenames

	def get_all_calib_filenames(self):
		filenames = []

		for file in os.listdir(self.get_calib_param_path()):
			if file.endswith(".txt"):
				filenames.append(file)

		return filenames

	def get_all_test_filenames(self):
		filenames = []

		for file in os.listdir(self.get_test_param_path()):
			if file.endswith(".txt"):
				filenames.append(file)

		return filenames

	def get_all_positioner_physics_filenames(self):
		filenames = []

		for file in os.listdir(self.get_positioner_physics_path()):
			if file.endswith(".txt"):
				filenames.append(file)

		return filenames

	def get_all_positioner_requirements_filenames(self):
		filenames = []

		for file in os.listdir(self.get_positioner_requirements_path()):
			if file.endswith(".txt"):
				filenames.append(file)

		return filenames

	def save_positioners_model(self, testBench):
		for positioner in testBench.positioners:
			filePath = os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultFolder,self.positionerFolderPrefix+'_'+str(positioner.ID))
			fileName = self.positionerModelFile+'_'+str(positioner.ID)+self.positionerModelExtension
			if self.overwritePositionerModel or not os.path.exists(os.path.join(filePath,fileName)):
				positioner.model.save(filePath, fileName)

	def load_positioners_model(self, testBench):
		for positioner in testBench.positioners:
			filePath = os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultFolder,self.positionerFolderPrefix+'_'+str(positioner.ID))
			fileName = os.path.join(filePath,self.positionerModelFile+'_'+str(positioner.ID)+self.positionerModelExtension)
			if os.path.exists(fileName):
				positioner.model.load(fileName)

	def load_calib_results(self, calibResults, positionerIDs, lifetimeLoop = 0):
		if len(calibResults) is not len(positionerIDs):
			raise errors.Error("Calibration result container has the wrong length") from None

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder)

		i = 0
		for positionerID in positionerIDs:
			if self.resultsLoadingFolder == DEFINES.CONFIG_LOAD_LATEST_RESULT:
				resultPath = self.get_latest_positioner_folder(positionerID)
				if resultPath == '':
					raise errors.IOError(f'Positioner {positionerID:04d} results folder not found') from None
			else:
				resultPath = self.resultsLoadingFolder

			resultPath = os.path.join(	filePath,\
										self.positionerFolderPrefix+'_'+str(positionerID),\
										resultPath)

			if self.check_folder_is_lifetime(resultPath):
				resultPath = os.path.join(	resultPath,\
											self.lifetimeIterationFolderName+'_'+str(lifetimeLoop+1))

			try:
				calibResults[i].load(os.path.join(	resultPath,\
													self.calibrationResultsFile+self.calibrationResultsFileExt))
			except errors.IOError as e:
				log.message(DEFINES.LOG_MESSAGE_PRIORITY_ERROR,0,str(e))
				raise errors.IOError(f'Positioner {positionerID:04d} calibration results loading failed') from None

			i += 1

	def save_calib_results(self, calibResults):
		for currentResult in calibResults:
			filePath = self.get_current_positioner_folder(currentResult.positionerID)
			fileName = self.calibrationResultsFile+self.calibrationResultsFileExt

			currentResult.save(filePath, fileName)

	def load_test_results(self, testResults, positionerIDs, lifetimeLoop = 0):
		if len(testResults) is not len(positionerIDs):
			raise errors.Error("Test result container has the wrong length") from None

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder)

		i = 0
		for positionerID in positionerIDs:
			if self.resultsLoadingFolder == DEFINES.CONFIG_LOAD_LATEST_RESULT:
				resultPath = self.get_latest_positioner_folder(positionerID)
				if resultPath == '':
					raise errors.IOError(f'Positioner {positionerID:04d} results folder not found') from None
			else:
				resultPath = self.resultsLoadingFolder

			resultPath = os.path.join(	filePath,\
										self.positionerFolderPrefix+'_'+str(positionerID),\
										resultPath)

			if self.check_folder_is_lifetime(resultPath):
				resultPath = os.path.join(	resultPath,\
											self.lifetimeIterationFolderName+'_'+str(lifetimeLoop+1))

			try:
				testResults[i].load(os.path.join(	resultPath,\
													self.testResultsFile+self.testResultsFileExt))
			except errors.IOError:
				log.message(DEFINES.LOG_MESSAGE_PRIORITY_ERROR,0,str(e))
				raise errors.IOError(f'Positioner {positionerID:04d} test results loading failed') from None

			i += 1

	def save_test_results(self, testResults):
		for currentResult in testResults:
			filePath = self.get_current_positioner_folder(currentResult.positionerID)
			fileName = self.testResultsFile+self.testResultsFileExt

			currentResult.save(filePath, fileName)

	def get_current_figure_folder(self, positionerID):

		filePath = os.path.join(	self.get_current_positioner_folder(positionerID),\
									self.figureFolder)

		return filePath

	def get_current_overview_folder(self):

		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.overviewsFolder)
		return filePath

	def get_overwiew_filename(self, positionerID):

		overviewFile = self.currentProjectTime+'_'+self.positionerFolderPrefix+'_'+str(positionerID)

		if self.positionerFolderSuffix is not '':
			overviewFile += '_'+self.positionerFolderSuffix
		if self.nbTestingLoops > 1:
			overviewFile += f'_{self.lifetimeSuffix}_{self.currentLifetimeIteration+1}'

		return overviewFile

	def get_current_positioner_folder(self, positionerID, includeLifetimeIteration = True):
	
		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder,\
									self.positionerFolderPrefix+'_'+str(positionerID),\
									self.currentProjectTime)

		if self.positionerFolderSuffix is not '':
			filePath += '_'+self.positionerFolderSuffix
		
		if self.nbTestingLoops > 1:
			filePath += '_'+self.lifetimeSuffix

			if includeLifetimeIteration:
				filePath = os.path.join(	filePath,\
											self.lifetimeIterationFolderName+'_'+str(self.currentLifetimeIteration+1))

		return filePath

	#get the last run done with the positioner, but excluding the current run
	def get_latest_positioner_folder(self, positionerID):
		#go to the project folder of the positioner
		filePath = os.path.join(	self.generalProjectFolder,\
									self.currentProjectFolder,\
									self.resultFolder,\
									self.positionerFolderPrefix+'_'+str(positionerID))

		#list all the runs performed with this positioner (get all the folder names)
		availableResultsFolders = [folderName for folderName in os.listdir(filePath) if os.path.isdir(os.path.join(filePath, folderName))]
		nbAvailableFolders = len(availableResultsFolders)

		if nbAvailableFolders < 1:
			return ''

		#extract the time out of the folder name
		availableResultsTimes = []
		for i in range(0,nbAvailableFolders):
			availableResultsTimes.append(availableResultsFolders[i].split('_')[0])
			if availableResultsTimes[-1] == self.currentProjectTime:
				del availableResultsTimes[-1]
			
		#get the latest project
		latestProjectIndex = max(range(0, len(availableResultsTimes)), key = lambda i: availableResultsTimes[i])

		return availableResultsFolders[latestProjectIndex]

	def check_folder_is_lifetime(self, folderPath):
		#check if the specified folder contains lifetime iterations or not
		subfolders = [folderName for folderName in os.listdir(folderPath) if os.path.isdir(os.path.join(folderPath, folderName))]
		nbSubfolders = len(subfolders)

		for i in range(0, nbSubfolders):
			subfolderPrefix = subfolders[i].split('_')[0]
			if subfolderPrefix == self.lifetimeIterationFolderName:
				return True

		return False

	def save_QC_result(self, calibResults = [], testResults = []):
		#Open the file
		if os.path.isfile(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile)):
			#The file exists in the project
			if os.path.isfile(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave)):
				#autosave exists from a previous failed save. load from it and delete it
				wb = openpyxl.load_workbook(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))
				os.remove(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))
			else:
				wb = openpyxl.load_workbook(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile))

		elif os.path.isfile(os.path.join(self.generalConfigFolder, self.configFolder, self.resultsOverviewTemplateFile)):
			#The template exists
			wb = openpyxl.load_workbook(os.path.join(self.generalConfigFolder, self.configFolder, self.resultsOverviewTemplateFile))
		else:
			raise errors.IOError('QC file: Neither the file nor the template do exist') #Neither the file nor the template do exist

		#Write results to the file
		ws1 = wb["Results"]

		if calibResults is not []:
			nbSlots = len(calibResults)
		else:
			nbSlots = len(testResults)

		for slot in range(0, nbSlots):
			QCpassed = False
			repeatabilityChecked = False
			hysteresisChecked = False
			writeLine = None
			if calibResults is not [] and slot < len(calibResults):
				alphaLength = calibResults[slot].mesAlphaLength[-1]
				betaLength = calibResults[slot].mesBetaLength[-1]
				RMSModelFit = calibResults[slot].mesRMSModelFit[-1]
				RMSRepeatability = calibResults[slot].mesRMSRepeatability[-1]
				maxHysteresis = calibResults[slot].mesMaxHysteresis[-1]
				maxNonLinearity = calibResults[slot].mesMaxNL[-1]
				maxNonLinDerivative = calibResults[slot].mesMaxNLDerivative[-1]
				RMSalignmentError = calibResults[slot].mesRMSAlignmentError[-1]
				maxAlignmentError = calibResults[slot].mesMaxAlignmentError[-1]
				roundnessDeviation = calibResults[slot].mesMaxRoundnessError[-1]
			
				alphaLengthPassed = abs(alphaLength-calibResults[slot].requirements.nominalAlphaLength) <= calibResults[slot].requirements.maxAlphaLengthDeviation
				betaLengthPassed = abs(betaLength-calibResults[slot].requirements.nominalBetaLength) <= calibResults[slot].requirements.maxBetaLengthDeviation
				RMSModelFitPassed = RMSModelFit <= calibResults[slot].requirements.maxPosError
				RMSRepeatabilityPassed = RMSRepeatability <= calibResults[slot].requirements.rmsPosRepeatability
				maxHysteresisPassed = maxHysteresis <= calibResults[slot].requirements.maxHysteresis
				maxNonLinearityPassed = maxNonLinearity <= calibResults[slot].requirements.maxNonLinearity
				maxNonLinDerivativePassed = maxNonLinDerivative <= calibResults[slot].requirements.maxNonLinearityDerivative
				roundnessDeviationPassed = roundnessDeviation <= calibResults[slot].requirements.maxRoundnessDeviation

				QCpassed = 	alphaLengthPassed and\
							betaLengthPassed and\
							maxNonLinearityPassed and\
							maxNonLinDerivativePassed and\
							roundnessDeviationPassed

				if not np.isnan(RMSRepeatability):
					QCpassed = QCpassed and RMSRepeatabilityPassed
					repeatabilityChecked = True

				if not np.isnan(maxHysteresis):
					QCpassed = QCpassed and maxHysteresisPassed
					hysteresisChecked = True
				else:
					QCpassed = False

				fontPassed = openpyxl.styles.Font(color = "008000")
				fontFailed = openpyxl.styles.Font(color = "FF0000")

				#get line to write. Either the first writable line or the line matching the ID
				i = 2
				while ws1.cell(row = i, column = 1).value is not None and not (ws1.cell(row = i, column = 1).value == calibResults[slot].positionerID):
					i += 1

				writeLine = i

				ws1.cell(row = writeLine, column = 1, value = calibResults[slot].positionerID) #A: ID
				ws1.cell(row = writeLine, column = 3, value = calibResults[slot].config.currentProjectTime) #C: Calib time
				ws1.cell(row = writeLine, column = 5, value = calibResults[slot].testBenchName) #E: Bench
				ws1.cell(row = writeLine, column = 6, value = int(calibResults[slot].slotID)) #F: Slot ID

				currentCell = ws1.cell(row = writeLine, column = 7, value = alphaLength) #G: Alpha length
				if alphaLengthPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 8, value = betaLength) #H: Beta length
				if betaLengthPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 9, value = RMSModelFit) #I: Model fit
				if RMSModelFitPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 10, value = RMSRepeatability) #J: Repeatability
				if RMSRepeatabilityPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 11, value = maxHysteresis) #K: Hysteresis
				if maxHysteresisPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 12, value = maxNonLinearity) #L: NL
				if maxNonLinearityPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 13, value = maxNonLinDerivative) #M: NL derivative
				if maxNonLinDerivativePassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 14, value = roundnessDeviation) #N: Roundness
				if roundnessDeviationPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

			if testResults is not [] and slot < len(testResults):
				nbRepetitions, nbTargets, maxNbMoves, nbDims = testResults[slot].targets.shape
				nbPoints = nbTargets*nbRepetitions

				RMSErrorFirstMove 			= testResults[slot].mesRMSError1stMove[-1]
				RMSRepeatabilityFirstMove 	= testResults[slot].mesRMSRepeatability1stMove[-1]
				targetConvergeance 			= testResults[slot].mesTargetConvergeance[-1][-1]
				maxNbMoves 					= testResults[slot].mesMaxNbMoves[-1]
				
				RMSErrorFirstMovePassed			= RMSErrorFirstMove <= testResults[slot].requirements.maxPosError
				RMSRepeatabilityFirstMovePassed = RMSRepeatabilityFirstMove <= testResults[slot].requirements.rmsPosRepeatability
				targetConvergeancePassed 		= targetConvergeance >= testResults[slot].requirements.targetConvergeance
				maxNbMovesPassed 				= maxNbMoves <= testResults[slot].requirements.maxNbMoves

				#get line to write. Either the first writable line or the line matching the ID
				if writeLine is None:
					i = 2
					while ws1.cell(row = i, column = 1).value is not None and not (ws1.cell(row = i, column = 1).value == testResults[slot].positionerID):
						i += 1

					writeLine = i

				currentCell = ws1.cell(row = writeLine, column = 4, value = testResults[slot].config.currentProjectTime) #D: Test time
				
				currentCell = ws1.cell(row = writeLine, column = 15, value = RMSErrorFirstMove) #O: Test RMS error
				if RMSErrorFirstMovePassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 16, value = RMSRepeatabilityFirstMove) #P: Test RMS repeatability
				if RMSRepeatabilityFirstMovePassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 17, value = targetConvergeance) #Q: Test convergeance
				if targetConvergeancePassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				currentCell = ws1.cell(row = writeLine, column = 18, value = maxNbMoves) #R: Test max moves
				if maxNbMovesPassed:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

				if not repeatabilityChecked:
					if not np.isnan(RMSRepeatabilityFirstMove):
						QCpassed = QCpassed and RMSRepeatabilityFirstMovePassed
					else : #Fail the result if the repeatability was not checked
						QCpassed = False

			if QCpassed and repeatabilityChecked and hysteresisChecked:
				QCresult = 'PASSED'
			else:
				QCresult = 'FAILED'

			if writeLine is not None:
				currentCell = ws1.cell(row = writeLine, column = 2, value = QCresult) #B: QA result
				if QCpassed and repeatabilityChecked and hysteresisChecked:
					currentCell.font = fontPassed
				else:
					currentCell.font = fontFailed

		#save the file
		try:
			wb.save(os.path.join(self.generalProjectFolder,self.currentProjectFolder,self.resultsOverviewFile))
		except PermissionError: #If the file is already opened, autosave a copy in the template folder
			wb.save(os.path.join(self.generalConfigFolder, self.configFolder,self.resultsOverviewAutosave))
		
def main():
	config = Config()
	config.load(config.get_config_fileName())
	print(config.get_latest_positioner_folder(101))
	print(config.check_folder_is_lifetime(config.get_latest_positioner_folder(101)))

if __name__ == '__main__':
	main()


